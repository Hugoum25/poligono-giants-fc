import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DIRECTORY = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(DIRECTORY, 'src', 'data')
EXCEL_PATH = os.path.join(DIRECTORY, 'Base_de_Datos_Giants.xlsx')
LAST_SYNC_FILE = os.path.join(DIRECTORY, '.last_excel_sync')

HEADER_FILL = PatternFill(start_color="FF2A85", end_color="FF2A85", fill_type="solid") # Rosa Giants
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT  = Font(name="Calibri", size=14, bold=True, color="161824")
BORDER_THIN = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD')
)

def export_all_to_excel():
    wb = openpyxl.Workbook()
    # Eliminar hoja por defecto
    default_sheet = wb.active
    
    # -------------------------------------------------------------
    # 1. JUGADORES
    # -------------------------------------------------------------
    players_file = os.path.join(DATA_DIR, 'players.json')
    if os.path.exists(players_file):
        with open(players_file, 'r', encoding='utf-8') as f:
            players = json.load(f)
        
        ws_players = wb.create_sheet(title="Jugadores")
        headers = ["ID", "Nombre", "Dorsal", "Posición", "Categoría", "Edad", "PJ", "Goles", "Asistencias", "MVP", "T. Amarillas", "T. Rojas", "T. Azules", "Victorias", "Empates", "Derrotas"]
        ws_players.append(headers)
        
        for p in players:
            stats = p.get('stats', {})
            info = p.get('info', {})
            row = [
                p.get('id'),
                p.get('name'),
                p.get('number'),
                p.get('position'),
                p.get('category'),
                info.get('age', ''),
                stats.get('matches', 0),
                stats.get('goals', 0),
                stats.get('assists', 0),
                stats.get('mvp', 0),
                stats.get('yellowCards', 0),
                stats.get('redCards', 0),
                stats.get('blueCards', 0),
                stats.get('wins', 0),
                stats.get('draws', 0),
                stats.get('losses', 0)
            ]
            ws_players.append(row)
        style_sheet(ws_players)

    # -------------------------------------------------------------
    # 2. PARTIDOS
    # -------------------------------------------------------------
    matches_file = os.path.join(DATA_DIR, 'matches.json')
    if os.path.exists(matches_file):
        with open(matches_file, 'r', encoding='utf-8') as f:
            matches = json.load(f)
        
        ws_matches = wb.create_sheet(title="Partidos")
        headers = ["ID", "Tipo", "Competición", "Rival", "Fecha", "Estadio", "Local/Visitante", "Goles Giants", "Goles Rival", "Ganado"]
        ws_matches.append(headers)
        
        for m in matches:
            row = [
                m.get('id'),
                m.get('type'),
                m.get('competition'),
                m.get('opponent'),
                m.get('date'),
                m.get('stadium', 'Campo Municipal La Camocha (Gijón)'),
                "Local" if m.get('isHome') else "Visitante",
                m.get('goalsGiants', ''),
                m.get('goalsOpponent', ''),
                m.get('isWin', '')
            ]
            ws_matches.append(row)
        style_sheet(ws_matches)

    # -------------------------------------------------------------
    # 3. NOTICIAS
    # -------------------------------------------------------------
    news_file = os.path.join(DATA_DIR, 'news.json')
    if os.path.exists(news_file):
        with open(news_file, 'r', encoding='utf-8') as f:
            news = json.load(f)
        
        ws_news = wb.create_sheet(title="Noticias")
        headers = ["ID", "Título", "Fecha", "Categoría", "Destacada", "Resumen", "Imagen"]
        ws_news.append(headers)
        
        for n in news:
            row = [
                n.get('id'),
                n.get('title'),
                n.get('date'),
                n.get('category'),
                "Sí" if n.get('featured') else "No",
                n.get('excerpt'),
                n.get('image')
            ]
            ws_news.append(row)
        style_sheet(ws_news)

    # -------------------------------------------------------------
    # 4. MULTIMEDIA
    # -------------------------------------------------------------
    media_file = os.path.join(DATA_DIR, 'media.json')
    if os.path.exists(media_file):
        with open(media_file, 'r', encoding='utf-8') as f:
            media = json.load(f)
        
        ws_media = wb.create_sheet(title="Multimedia")
        headers = ["ID", "Tipo", "Título", "Categoría", "URL / Archivo", "Vistas"]
        ws_media.append(headers)
        
        for item in media:
            row = [
                item.get('id'),
                item.get('type'),
                item.get('title'),
                item.get('category'),
                item.get('videoUrl') or item.get('image'),
                item.get('views', 0)
            ]
            ws_media.append(row)
        style_sheet(ws_media)

    # Eliminar hoja residual si existe
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
        
    wb.save(EXCEL_PATH)
    # Registrar mtime para evitar auto-import tras exportación
    try:
        with open(LAST_SYNC_FILE, 'w') as f:
            f.write(str(os.path.getmtime(EXCEL_PATH)))
    except: pass

    print(f"[Excel DB] Exportación completada con éxito en: {EXCEL_PATH}")
    return EXCEL_PATH

def import_excel_to_json():
    if not os.path.exists(EXCEL_PATH):
        return False
    
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    # 1. JUGADORES
    if "Jugadores" in wb.sheetnames:
        ws = wb["Jugadores"]
        players = []
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            for r in rows[1:]:
                if r[0] is None: continue
                players.append({
                    "id": int(r[0]),
                    "name": str(r[1] or ""),
                    "number": int(r[2]) if r[2] is not None else 0,
                    "position": str(r[3] or ""),
                    "category": str(r[4] or "jugadores"),
                    "stats": {
                        "matches": int(r[6]) if r[6] is not None else 0,
                        "goals": int(r[7]) if r[7] is not None else 0,
                        "assists": int(r[8]) if r[8] is not None else 0,
                        "mvp": int(r[9]) if len(r) > 9 and r[9] is not None else 0,
                        "yellowCards": int(r[10]) if len(r) > 10 and r[10] is not None else 0,
                        "redCards": int(r[11]) if len(r) > 11 and r[11] is not None else 0,
                        "blueCards": int(r[12]) if len(r) > 12 and r[12] is not None else 0,
                        "wins": int(r[13]) if len(r) > 13 and r[13] is not None else 0,
                        "draws": int(r[14]) if len(r) > 14 and r[14] is not None else 0,
                        "losses": int(r[15]) if len(r) > 15 and r[15] is not None else 0
                    },
                    "info": {
                        "age": int(r[5]) if r[5] is not None and str(r[5]).isdigit() else 24
                    }
                })
            with open(os.path.join(DATA_DIR, 'players.json'), 'w', encoding='utf-8') as f:
                json.dump(players, f, ensure_ascii=False, indent=2)

    # 2. PARTIDOS
    if "Partidos" in wb.sheetnames:
        ws = wb["Partidos"]
        matches = []
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            for r in rows[1:]:
                if r[0] is None: continue
                matches.append({
                    "id": int(r[0]),
                    "type": str(r[1] or "past"),
                    "competition": str(r[2] or ""),
                    "opponent": str(r[3] or ""),
                    "date": str(r[4] or ""),
                    "stadium": str(r[5] or "Campo Municipal La Camocha (Gijón)"),
                    "isHome": True if str(r[6]).lower() in ['local', 'true', 'sí', 'si', '1'] else False,
                    "goalsGiants": int(r[7]) if r[7] is not None and str(r[7]).isdigit() else None,
                    "goalsOpponent": int(r[8]) if r[8] is not None and str(r[8]).isdigit() else None,
                    "isWin": True if str(r[9]).lower() in ['true', 'sí', 'si', '1'] else False
                })
            with open(os.path.join(DATA_DIR, 'matches.json'), 'w', encoding='utf-8') as f:
                json.dump(matches, f, ensure_ascii=False, indent=2)

    # 3. NOTICIAS
    if "Noticias" in wb.sheetnames:
        ws = wb["Noticias"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            existing_news = []
            news_path = os.path.join(DATA_DIR, 'news.json')
            if os.path.exists(news_path):
                try:
                    with open(news_path, 'r', encoding='utf-8') as f:
                        existing_news = json.load(f)
                except: pass
            
            news_dict = {n.get('id'): n for n in existing_news}
            for r in rows[1:]:
                if r[0] is None: continue
                nid = int(r[0])
                n_obj = news_dict.get(nid, {})
                n_obj['id'] = nid
                n_obj['title'] = str(r[1] or "")
                n_obj['date'] = str(r[2] or "")
                n_obj['category'] = str(r[3] or "club")
                n_obj['featured'] = True if str(r[4]).lower() in ['sí', 'si', 'true', '1'] else False
                n_obj['excerpt'] = str(r[5] or "")
                if r[6]: n_obj['image'] = str(r[6])
                news_dict[nid] = n_obj
            
            updated_news = list(news_dict.values())
            with open(news_path, 'w', encoding='utf-8') as f:
                json.dump(updated_news, f, ensure_ascii=False, indent=2)

    # 4. MULTIMEDIA
    if "Multimedia" in wb.sheetnames:
        ws = wb["Multimedia"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            media = []
            for r in rows[1:]:
                if r[0] is None: continue
                item = {
                    "id": int(r[0]),
                    "type": str(r[1] or "photo"),
                    "title": str(r[2] or ""),
                    "category": str(r[3] or "general"),
                    "views": int(r[5]) if r[5] is not None and str(r[5]).isdigit() else 0
                }
                if item["type"] == "video":
                    item["videoUrl"] = str(r[4] or "")
                else:
                    item["image"] = str(r[4] or "")
                media.append(item)
            with open(os.path.join(DATA_DIR, 'media.json'), 'w', encoding='utf-8') as f:
                json.dump(media, f, ensure_ascii=False, indent=2)

    # Actualizar marca mtime
    try:
        with open(LAST_SYNC_FILE, 'w') as f:
            f.write(str(os.path.getmtime(EXCEL_PATH)))
    except: pass

    print("[Excel DB] Importación desde Excel completada exitosamente.")
    return True

def sync_excel_if_modified():
    if not os.path.exists(EXCEL_PATH):
        return False
    excel_mtime = os.path.getmtime(EXCEL_PATH)
    last_sync = 0
    if os.path.exists(LAST_SYNC_FILE):
        try:
            with open(LAST_SYNC_FILE, 'r') as f:
                last_sync = float(f.read().strip())
        except: pass
    
    if excel_mtime > (last_sync + 1.5):
        print("[Excel DB] Detectada modificación en Excel. Sincronizando cambios a la web...")
        try:
            import_excel_to_json()
            return True
        except Exception as e:
            print(f"[Excel DB] Error importando cambios de Excel: {e}")
            return False
    return False

def style_sheet(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 1:
                cell.border = BORDER_THIN
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

if __name__ == "__main__":
    export_all_to_excel()
